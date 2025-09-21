"""
File Processor Service for extracting content from various file types
"""

import os
import logging
import mimetypes
from typing import Dict, List, Any, Optional
import PyPDF2
import openpyxl
import docx
from PIL import Image
import pytesseract
import requests
from io import BytesIO
import json

logger = logging.getLogger(__name__)

class FileProcessor:
    def __init__(self):
        self.supported_types = {
            'pdf': self._extract_pdf_content,
            'docx': self._extract_docx_content,
            'doc': self._extract_docx_content,
            'xlsx': self._extract_excel_content,
            'xls': self._extract_excel_content,
            'csv': self._extract_csv_content,
            'txt': self._extract_text_content,
            'png': self._extract_image_content,
            'jpg': self._extract_image_content,
            'jpeg': self._extract_image_content,
            'mp3': self._extract_audio_content,
            'wav': self._extract_audio_content,
            'mp4': self._extract_video_content,
            'avi': self._extract_video_content,
            'mov': self._extract_video_content
        }
    
    def extract_content(self, file_path: str, file_type: str) -> str:
        """
        Extract content from a file based on its type
        """
        try:
            # Get file extension
            file_ext = file_type.lower()
            
            if file_ext in self.supported_types:
                return self.supported_types[file_ext](file_path)
            else:
                logger.warning(f"Unsupported file type: {file_ext}")
                return f"Unsupported file type: {file_ext}"
                
        except Exception as e:
            logger.error(f"Error extracting content from {file_path}: {e}")
            return f"Error extracting content: {str(e)}"
    
    def _extract_pdf_content(self, file_path: str) -> str:
        """Extract text content from PDF files"""
        try:
            content = []
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    content.append(page.extract_text())
            
            return '\n'.join(content)
        except Exception as e:
            logger.error(f"Error extracting PDF content: {e}")
            return f"Error extracting PDF content: {str(e)}"
    
    def _extract_docx_content(self, file_path: str) -> str:
        """Extract text content from DOCX files"""
        try:
            doc = docx.Document(file_path)
            content = []
            for paragraph in doc.paragraphs:
                content.append(paragraph.text)
            
            return '\n'.join(content)
        except Exception as e:
            logger.error(f"Error extracting DOCX content: {e}")
            return f"Error extracting DOCX content: {str(e)}"
    
    def _extract_excel_content(self, file_path: str) -> str:
        """Extract content from Excel files"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            content = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                content.append(f"Sheet: {sheet_name}")
                
                # Extract data from each cell
                for row in sheet.iter_rows(values_only=True):
                    row_data = []
                    for cell in row:
                        if cell is not None:
                            row_data.append(str(cell))
                    if row_data:
                        content.append('\t'.join(row_data))
                
                content.append('')  # Empty line between sheets
            
            return '\n'.join(content)
        except Exception as e:
            logger.error(f"Error extracting Excel content: {e}")
            return f"Error extracting Excel content: {str(e)}"
    
    def _extract_csv_content(self, file_path: str) -> str:
        """Extract content from CSV files"""
        try:
            import csv
            content = []
            with open(file_path, 'r', encoding='utf-8') as file:
                csv_reader = csv.reader(file)
                for row in csv_reader:
                    content.append('\t'.join(row))
            
            return '\n'.join(content)
        except Exception as e:
            logger.error(f"Error extracting CSV content: {e}")
            return f"Error extracting CSV content: {str(e)}"
    
    def _extract_text_content(self, file_path: str) -> str:
        """Extract content from text files"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except Exception as e:
            logger.error(f"Error extracting text content: {e}")
            return f"Error extracting text content: {str(e)}"
    
    def _extract_image_content(self, file_path: str) -> str:
        """Extract text content from images using OCR"""
        try:
            # Check if tesseract is available
            try:
                pytesseract.get_tesseract_version()
            except:
                return f"OCR not available for image: {os.path.basename(file_path)}"
            
            image = Image.open(file_path)
            text = pytesseract.image_to_string(image)
            return text
        except Exception as e:
            logger.error(f"Error extracting image content: {e}")
            return f"Error extracting image content: {str(e)}"
    
    def _extract_audio_content(self, file_path: str) -> str:
        """Extract content from audio files (placeholder)"""
        # This would require speech-to-text services like Google Speech-to-Text
        # For now, return a placeholder
        return f"Audio file detected: {os.path.basename(file_path)}. Speech-to-text processing not implemented yet."
    
    def _extract_video_content(self, file_path: str) -> str:
        """Extract content from video files (placeholder)"""
        # This would require video processing and speech-to-text
        # For now, return a placeholder
        return f"Video file detected: {os.path.basename(file_path)}. Video processing not implemented yet."
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get basic information about a file"""
        try:
            stat = os.stat(file_path)
            mime_type, _ = mimetypes.guess_type(file_path)
            
            return {
                'filename': os.path.basename(file_path),
                'size': stat.st_size,
                'mime_type': mime_type,
                'modified_time': stat.st_mtime
            }
        except Exception as e:
            logger.error(f"Error getting file info for {file_path}: {e}")
            return {
                'filename': os.path.basename(file_path),
                'size': 0,
                'mime_type': 'unknown',
                'modified_time': 0
            }
    
    def process_uploaded_files(self, uploaded_assets: List[Dict[str, Any]], base_upload_path: str) -> Dict[str, str]:
        """
        Process all uploaded files and extract their content
        """
        file_contents = {}
        
        for asset in uploaded_assets:
            try:
                file_type = asset.get('type', '')
                filename = asset.get('filename', '')
                file_path = os.path.join(base_upload_path, filename)
                
                if not os.path.exists(file_path):
                    logger.warning(f"File not found: {file_path}")
                    continue
                
                # Extract content
                content = self.extract_content(file_path, file_type)
                file_contents[file_type] = content
                
                logger.info(f"Extracted content from {filename} ({file_type})")
                
            except Exception as e:
                logger.error(f"Error processing file {asset.get('filename', 'unknown')}: {e}")
                continue
        
        return file_contents

# Global file processor instance
file_processor = FileProcessor()
